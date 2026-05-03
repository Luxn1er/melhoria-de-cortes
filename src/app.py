"""Aplicação principal — janela MRX com sidebar, relatório, visualização e exportação."""

from __future__ import annotations
import os
import queue
import re
import sys
import threading
import time
from typing import Dict, List, Optional, Tuple

import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox

from src.models import Bobina, Jumbo, Puxada
from src.optimizer import OtimizadorProducao
from src.database import MRXDatabase
from src.policy import RefilePolicy
from src.helpers import (
    formatar_lista_larguras,
    montar_larguras_puxada,
    normalizar_residuais,
    pendentes_apos_base,
    aplicar_layout_na_regua,
    expandir_base_tuple,
    agrupamento_base_sobras,
    intercalar_eixos,
    calcular_facas_puxada,
)
from src.ui.report import renderizar_relatorio
from src.ui.canvas_viz import desenhar_puxada
from src.ui.sobras_window import JanelaSobras


# ===========================================================================
# CONFIGURAÇÃO VISUAL
# ===========================================================================

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")


def _obter_base_dir() -> str:
    """Retorna o diretorio real do projeto, mesmo quando embutido (.exe)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(os.path.abspath(sys.argv[0]))
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# ===========================================================================
# APP PRINCIPAL
# ===========================================================================

class AppMRX(ctk.CTk):
    """Janela principal do Otimizador de Corte Industrial."""

    def __init__(self) -> None:
        super().__init__()
        self.title("MRX - Otimizador de Corte Industrial v2.0")
        self.geometry("1300x820")

        # Estado
        self.itens_entrada: List[Tuple[int, int]] = []
        self.estoque_atividades: dict[int, int] = {}
        self.plano_atual: List[Puxada] = []
        self.residuais_atuais: List[Tuple[int, int]] = []
        self.ultima_execucao_id: Optional[int] = None
        self.medida_inicial_atual: int = 795
        self._historico_adicoes: List[List[Tuple[int, int]]] = []  # lotes cronologicos
        self._progress_queue: queue.Queue[Tuple[float, str]] = queue.Queue()

        # Banco de dados
        _base = _obter_base_dir()
        data_dir = os.path.join(_base, "ProduçãoAlt")
        self.db = MRXDatabase(os.path.join(data_dir, "mrx_otimizador.sqlite3"))

        # Layout
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self._build_sidebar()
        self._build_main_panel()

        self._carregar_dados_iniciais()
        self._atualizar_texto_estoque()

    # ---- sidebar -----------------------------------------------------------

    def _build_sidebar(self) -> None:
        self.sidebar = ctk.CTkFrame(self, width=300)
        self.sidebar.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=10, pady=10)

        ctk.CTkLabel(
            self.sidebar, text="⚙️ CONFIGURAÇÃO", font=("Roboto", 18, "bold")
        ).pack(pady=20)

        self.ent_jumbo = ctk.CTkEntry(self.sidebar, placeholder_text="Jumbo (mm)")
        self.ent_jumbo.insert(0, "1565")
        self.ent_jumbo.pack(pady=10, padx=20)

        self.ent_medida_inicial = ctk.CTkEntry(
            self.sidebar, placeholder_text="Med. inicial (mm)"
        )
        self.ent_medida_inicial.insert(0, str(self.medida_inicial_atual))
        self.ent_medida_inicial.pack(pady=10, padx=20)

        ctk.CTkLabel(self.sidebar, text="Adicionar Bobina:").pack(pady=(20, 5))
        self.ent_larg = ctk.CTkEntry(self.sidebar, placeholder_text="Largura (mm)")
        self.ent_larg.pack(pady=5, padx=20)
        self.ent_qtd = ctk.CTkEntry(self.sidebar, placeholder_text="Quantidade")
        self.ent_qtd.pack(pady=5, padx=20)

        ctk.CTkButton(
            self.sidebar, text="➕ Adicionar", command=self.add_bobina
        ).pack(pady=20, padx=20)
        ctk.CTkButton(
            self.sidebar, text="Importar Puxadas", command=self.abrir_importar_puxadas
        ).pack(pady=10, padx=20)
        ctk.CTkButton(
            self.sidebar, text="📤 Exportar Planilha", command=self.exportar_planilha
        ).pack(pady=10, padx=20)
        ctk.CTkButton(
            self.sidebar, text="↩️ Desfazer",
            command=self.desfazer,
        ).pack(pady=10, padx=20)
        ctk.CTkButton(
            self.sidebar, text="🗑️ Limpar Tudo",
            fg_color="#c0392b", command=self.limpar,
        ).pack(pady=10, padx=20)

        # Legenda
        ctk.CTkLabel(self.sidebar, text="", height=10).pack()
        ctk.CTkLabel(
            self.sidebar, text="LEGENDA DE REFILE", font=("Roboto", 12, "bold")
        ).pack(pady=(10, 4), padx=20)
        for txt in [
            "🟢 Faixa Primária: 10–15mm/lado",
            "🟡 Faixa Secundária: 15–25mm/lado",
            "🔴 Residual: bobinas sem padrão",
        ]:
            ctk.CTkLabel(self.sidebar, text=txt, font=("Roboto", 11), anchor="w"
                         ).pack(padx=20, anchor="w")

        # Progresso + status
        self.lbl_status_sidebar = ctk.CTkLabel(
            self.sidebar, text="", anchor="center",
            fg_color="transparent", font=("Roboto", 11),
        )
        self.progress_bar = ctk.CTkProgressBar(self.sidebar, mode="determinate")
        self.progress_bar.set(0)

    # ---- painel principal --------------------------------------------------

    def _build_main_panel(self) -> None:
        self.main = ctk.CTkFrame(self)
        self.main.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        self.main.grid_columnconfigure(0, weight=3)
        self.main.grid_columnconfigure(1, weight=2)
        self.main.grid_rowconfigure(0, weight=1)
        self.main.grid_rowconfigure(1, weight=0)

        # Frame central — estoque / relatório
        self.frame_central = ctk.CTkFrame(self.main, fg_color="transparent")
        self.frame_central.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
        self.frame_central.grid_rowconfigure(0, weight=0)
        self.frame_central.grid_rowconfigure(1, weight=1)
        self.frame_central.grid_columnconfigure(0, weight=1)

        self.lbl_estoque_titulo = ctk.CTkLabel(
            self.frame_central, text="ESTOQUE ATUAL:",
            font=("Roboto", 15, "bold"), anchor="w",
        )
        self.lbl_estoque_titulo.grid(row=0, column=0, sticky="nw", padx=4, pady=(4, 8))

        self.txt = ctk.CTkTextbox(
            self.frame_central, font=("Consolas", 13),
            border_width=0, fg_color="transparent",
        )
        self.txt.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)

        # Painel de visualização
        self.viz = ctk.CTkFrame(self.main)
        self.viz.grid(row=0, column=1, sticky="nsew")
        self.viz.grid_rowconfigure(2, weight=1)
        self.viz.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.viz, text="📊 Visualização da Puxada",
            font=("Roboto", 16, "bold"),
        ).grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))

        self.cmb_puxada = ctk.CTkOptionMenu(
            self.viz, values=["(sem puxadas)"], command=self._on_select_puxada,
        )
        self.cmb_puxada.grid(row=1, column=0, sticky="ew", padx=10, pady=10)

        self.canvas = tk.Canvas(
            self.viz, bg="#0f0f10", highlightthickness=1, highlightbackground="#2b2b2b"
        )
        self.canvas.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.canvas.bind("<Configure>", lambda _e: self._redesenhar_canvas())

        # Botão gerar
        self.btn_gerar = ctk.CTkButton(
            self.main, text="🚀 GERAR PLANO DE PRODUÇÃO",
            height=60, font=("Roboto", 18, "bold"),
            fg_color="#27ae60", command=self.processar,
        )
        self.btn_gerar.grid(row=1, column=0, columnspan=2, sticky="ew", padx=0, pady=(12, 0))

    # ---- ações da sidebar --------------------------------------------------

    def add_bobina(self) -> None:
        try:
            l, q = int(self.ent_larg.get()), int(self.ent_qtd.get())
            if l <= 0 or q <= 0:
                messagebox.showwarning("Aviso", "Largura e quantidade devem ser maiores que zero.")
                return
            self.estoque_atividades[l] = self.estoque_atividades.get(l, 0) + q
            self.itens_entrada = sorted(
                self.estoque_atividades.items(), key=lambda x: x[0], reverse=True
            )
            self.db.upsert_estoque(l, self.estoque_atividades[l])
            self._historico_adicoes.append([(l, q)])
            self.ent_larg.delete(0, "end")
            self.ent_qtd.delete(0, "end")
            self._atualizar_texto_estoque()
        except Exception as e:
            messagebox.showerror("Erro", f"Dados inválidos.\n{e}")

    def abrir_importar_puxadas(self) -> None:
        """Abre janela para colar varias bobinas de uma vez."""
        win = ctk.CTkToplevel(self)
        win.title("Importar Puxadas")
        win.geometry("620x520")
        win.resizable(True, True)
        win.transient(self)
        win.grab_set()
        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(win, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 8))
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="Cole as bobinas copiadas da planilha",
            font=("Roboto", 18, "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(
            header,
            text="Formatos aceitos: 1200 3, 1200;3, 1200x3 ou somente 1200 para quantidade 1.",
            font=("Roboto", 12),
            text_color="#A0A0A0",
            anchor="w",
            wraplength=560,
        ).grid(row=1, column=0, sticky="ew", pady=(6, 0))

        txt_import = ctk.CTkTextbox(win, font=("Consolas", 13))
        txt_import.grid(row=1, column=0, sticky="nsew", padx=16, pady=8)

        lbl_status = ctk.CTkLabel(
            win, text="", anchor="w", text_color="#FFC107", wraplength=560
        )
        lbl_status.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 8))

        footer = ctk.CTkFrame(win, fg_color="transparent")
        footer.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 16))
        footer.grid_columnconfigure(0, weight=1)
        footer.grid_columnconfigure(1, weight=1)

        def confirmar() -> None:
            conteudo = txt_import.get("1.0", "end")
            itens, avisos = self._parse_importacao_puxadas(conteudo)
            if not itens:
                lbl_status.configure(
                    text="Nenhuma linha valida encontrada. Cole pelo menos uma largura."
                )
                return

            agregados: Dict[int, int] = {}
            for largura, qtd in itens:
                agregados[largura] = agregados.get(largura, 0) + qtd

            lote = sorted(agregados.items(), key=lambda x: x[0], reverse=True)
            for largura, qtd in lote:
                self.estoque_atividades[largura] = (
                    self.estoque_atividades.get(largura, 0) + qtd
                )
                self.db.upsert_estoque(largura, self.estoque_atividades[largura])

            self._historico_adicoes.append(lote)
            self.itens_entrada = sorted(
                self.estoque_atividades.items(), key=lambda x: x[0], reverse=True
            )
            self._atualizar_texto_estoque()

            total_bobinas = sum(qtd for _, qtd in lote)
            msg = f"{total_bobinas} bobina(s) importada(s) em {len(lote)} medida(s)."
            if avisos:
                msg += f"\n{len(avisos)} linha(s) ignorada(s)."
            messagebox.showinfo("Importacao concluida", msg)
            win.destroy()

        ctk.CTkButton(
            footer, text="Importar", fg_color="#287d3c", command=confirmar
        ).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ctk.CTkButton(
            footer, text="Cancelar", fg_color="#b03a2e", command=win.destroy
        ).grid(row=0, column=1, sticky="ew", padx=(8, 0))

        txt_import.focus_set()

    @staticmethod
    def _parse_importacao_puxadas(conteudo: str) -> Tuple[List[Tuple[int, int]], List[str]]:
        """
        Interpreta linhas coladas de planilha/texto.

        Padrao principal: largura quantidade. Com uma unica largura, usa qtd=1.
        Tambem aceita notacao com "x", como 1200x3 ou 3x1200.
        """
        itens: List[Tuple[int, int]] = []
        avisos: List[str] = []

        for idx, raw in enumerate(conteudo.splitlines(), start=1):
            linha = raw.strip()
            if not linha:
                continue

            nums = [int(n) for n in re.findall(r"\d+", linha)]
            if not nums:
                avisos.append(f"Linha {idx}: sem numeros")
                continue

            pares: List[Tuple[int, int]] = []
            if "x" in linha.lower() and len(nums) >= 2:
                a, b = nums[0], nums[1]
                if a <= 100 and b > a:
                    pares.append((b, a))
                else:
                    pares.append((a, b))
            elif len(nums) == 1:
                pares.append((nums[0], 1))
            elif len(nums) % 2 == 0:
                pares.extend((nums[i], nums[i + 1]) for i in range(0, len(nums), 2))
            else:
                pares.append((nums[0], nums[1]))

            for largura, qtd in pares:
                if largura <= 0 or qtd <= 0:
                    avisos.append(f"Linha {idx}: largura/quantidade invalida")
                    continue
                itens.append((int(largura), int(qtd)))

        return itens, avisos

    def desfazer(self) -> None:
        """Remove a última bobina adicionada (desfaz)."""
        if not self._historico_adicoes:
            messagebox.showinfo("Info", "Nada para desfazer.")
            return
        lote = self._historico_adicoes.pop()
        # Remove do estoque
        for ult_larg, ult_qtd in lote:
            if ult_larg in self.estoque_atividades:
                self.estoque_atividades[ult_larg] -= ult_qtd
                if self.estoque_atividades[ult_larg] <= 0:
                    del self.estoque_atividades[ult_larg]
                else:
                    self.db.upsert_estoque(ult_larg, self.estoque_atividades[ult_larg])
        self.db.substituir_estoque(list(self.estoque_atividades.items()))
        self.itens_entrada = sorted(
            self.estoque_atividades.items(), key=lambda x: x[0], reverse=True
        )
        self._atualizar_texto_estoque()

    def limpar(self) -> None:
        try:
            self.itens_entrada = []
            self.estoque_atividades = {}
            self.plano_atual = []
            self.residuais_atuais = []
            self.ultima_execucao_id = None
            self._historico_adicoes = []
            self.txt.delete("1.0", "end")
            self.db.limpar_estoque()
            self._atualizar_texto_estoque()
            self._set_opcoes_puxada([])
            self._redesenhar_canvas()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao limpar.\n{e}")

    # ---- processamento -----------------------------------------------------

    def processar(self) -> None:
        if not self.itens_entrada:
            messagebox.showwarning(
                "Aviso", "Adicione bobinas ao estoque antes de gerar puxadas."
            )
            return
        try:
            jumbo_val = int(self.ent_jumbo.get())
            if jumbo_val <= 0:
                messagebox.showwarning("Aviso", "Jumbo deve ser maior que zero.")
                return
        except Exception as e:
            messagebox.showerror("Erro", f"Jumbo inválido.\n{e}")
            return
        try:
            medida_inicial_val = int(self.ent_medida_inicial.get())
            if medida_inicial_val <= 0:
                messagebox.showwarning(
                    "Aviso", "Medida inicial deve ser maior que zero."
                )
                return
            self.medida_inicial_atual = medida_inicial_val
        except Exception as e:
            messagebox.showerror("Erro", f"Medida inicial inválida.\n{e}")
            return

        self._iniciar_loading("Processando algoritmos...")

        def worker() -> None:
            try:
                time.sleep(0.12)
                jumbo = Jumbo(jumbo_val)
                otimizador = OtimizadorProducao(jumbo)
                for l, q in self.itens_entrada:
                    otimizador.adicionar_material(l, q)
                otimizador.rodar_otimizacao(
                    on_progress=lambda f, m: self._progress_queue.put((f, m))
                )
                self.after(
                    0,
                    lambda o=otimizador: self._finalizar_processamento(
                        jumbo,
                        o.plano, o.residuais,
                        o.abrir_janela_sobras,
                        o.sugestao_base_residuo,
                        o.sobra_residuo_mm,
                    ),
                )
            except Exception as e:
                self.after(0, lambda: self._falha_processamento(e))

        threading.Thread(target=worker, daemon=True).start()
        self._poll_progresso()

    def _reprocessar_estoque_atual(self, jumbo_mm: int) -> None:
        """Roda otimização novamente após confirmação manual de sobras."""
        if not self.estoque_atividades:
            messagebox.showinfo(
                "Puxada de sobras",
                "Composição confirmada e adicionada ao plano.\n"
                "Todas as sobras foram tratadas.",
            )
            return

        self._iniciar_loading("Recalculando plano com estoque atualizado...")

        def worker() -> None:
            try:
                time.sleep(0.08)
                jumbo = Jumbo(int(jumbo_mm))
                otimizador = OtimizadorProducao(jumbo)
                for l, q in sorted(
                    self.estoque_atividades.items(), key=lambda x: x[0], reverse=True
                ):
                    otimizador.adicionar_material(l, q)
                otimizador.rodar_otimizacao(
                    on_progress=lambda f, m: self._progress_queue.put((f, m))
                )
                self.after(
                    0,
                    lambda o=otimizador: self._finalizar_processamento(
                        jumbo,
                        self.plano_atual + o.plano,
                        o.residuais, o.abrir_janela_sobras,
                        o.sugestao_base_residuo,
                        o.sobra_residuo_mm,
                        o.refile_insuficiente_detectado,
                    ),
                )
            except Exception as e:
                self.after(0, lambda: self._falha_processamento(e))

        threading.Thread(target=worker, daemon=True).start()
        self._poll_progresso()

    def _falha_processamento(self, e: Exception) -> None:
        self._parar_loading()
        messagebox.showerror("Erro de Processamento", f"Verifique os dados:\n{e}")

    def _finalizar_processamento(
        self,
        jumbo: Jumbo,
        plano: List[Puxada],
        residuais: List[Tuple[int, int]],
        abrir_janela_sobras: bool = False,
        sugestao_base_residuo: Optional[List[int]] = None,
        sobra_residuo_mm: int = 0,
        refile_insuficiente_detectado: bool = False,
    ) -> None:
        try:
            self.plano_atual = plano
            self.residuais_atuais = residuais
            self.estoque_atividades = {int(w): int(q) for w, q in residuais}
            self.itens_entrada = sorted(
                self.estoque_atividades.items(), key=lambda x: x[0], reverse=True
            )
            self.db.substituir_estoque(residuais)
            self.ultima_execucao_id = self.db.salvar_execucao_puxadas(
                jumbo.largura_mm, plano
            )

            self._parar_loading()
            self._atualizar_texto_estoque()
            self._renderizar_relatorio(plano, residuais)
            self._set_opcoes_puxada(plano, residuais)

            if abrir_janela_sobras and residuais:
                self._tratar_residuais(
                    jumbo, residuais, sugestao_base_residuo, sobra_residuo_mm
                )
            elif refile_insuficiente_detectado and residuais:
                messagebox.showwarning(
                    "Refile Insuficiente",
                    "Não é possível finalizar a próxima puxada com as sobras atuais,\n"
                    "pois o melhor encaixe gera refile total menor que 20mm.\n"
                    "Ajuste o estoque para continuar.",
                )
            elif not residuais:
                messagebox.showinfo(
                    "Sucesso",
                    "Plano de produção gerado e salvo!\nTodas as bobinas foram alocadas.",
                )
        except Exception as e:
            self._parar_loading()
            messagebox.showerror("Erro", f"Falha ao finalizar puxadas:\n{e}")

    def _tratar_residuais(
        self,
        jumbo: Jumbo,
        residuais: List[Tuple[int, int]],
        base_sugestao: Optional[List[int]] = None,
        sobra_mm: Optional[int] = None,
    ) -> None:
        if not residuais:
            return
        JanelaSobras(self, jumbo, residuais, app_ref=self,
                      base_sugestao=base_sugestao, sobra_mm=sobra_mm)

    # ---- relatório ---------------------------------------------------------

    def _renderizar_relatorio(
        self, plano: List[Puxada], residuais: List[Tuple[int, int]]
    ) -> None:
        self.lbl_estoque_titulo.grid_remove()
        renderizar_relatorio(
            self.txt, plano, residuais, self.medida_inicial_atual
        )

    # ---- exportação --------------------------------------------------------

    def exportar_planilha(self) -> None:
        try:
            if not self.plano_atual:
                messagebox.showwarning("Aviso", "Gere puxadas antes de exportar.")
                return

            from collections import Counter
            from datetime import datetime
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            base_dir = _obter_base_dir()
            pasta = os.path.join(base_dir, "ProduçãoAlt")
            os.makedirs(pasta, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            caminho = os.path.join(pasta, f"Puxadas_MRX_{ts}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "Puxadas"
            ws_facas = wb.create_sheet("Facas")

            # Estilos
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
            peach_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            plano_exportacao = self._agrupar_puxadas_iguais_exportacao(
                self.plano_atual
            )

            for i, (p, repeticao_total) in enumerate(plano_exportacao, start=1):
                col = i
                # Bobinas do padrão (sem repetir pela repeticao)
                bobina_list: List[int] = [int(bob.largura) for bob in p.bobinas]

                # Cabeçalho com repetição total, agrupando puxadas iguais
                title = f"PUXADA {i:02d} ({repeticao_total}x)"
                self._escrever_celula(ws, row=1, col=col, value=title,
                                      font=header_font, fill=header_fill,
                                      alignment=header_center, border=thin_border)

                # Bobinas individuais, a partir da linha 2
                for j, texto in enumerate(bobina_list, start=2):
                    self._escrever_celula(ws, row=j, col=col, value=texto,
                                          alignment=center, border=thin_border)

                self._escrever_bloco_facas_excel(
                    ws_facas,
                    puxada=p,
                    indice=i,
                    repeticao_total=repeticao_total,
                    start_col=1 + (i - 1) * 7,
                    medida_inicial_mm=self.medida_inicial_atual,
                    header_font=header_font,
                    header_fill=header_fill,
                    yellow_fill=yellow_fill,
                    peach_fill=peach_fill,
                    center=center,
                    header_center=header_center,
                    thin_border=thin_border,
                )

            # Ajustar largura das colunas automaticamente
            for cell in ws[1]:
                col_letter = cell.column_letter
                ws.column_dimensions[col_letter].width = 18
            for col in range(1, ws_facas.max_column + 1):
                ws_facas.column_dimensions[
                    ws_facas.cell(row=1, column=col).column_letter
                ].width = 13

            ws.freeze_panes = "A1"
            ws_facas.freeze_panes = "A4"

            wb.save(caminho)

            messagebox.showinfo("Exportação concluída", f"Planilha salva em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar planilha:\n{e}")

    @staticmethod
    def _escrever_celula(ws, row, col, value, font=None, fill=None,
                         alignment=None, border=None):
        cel = ws.cell(row=row, column=col, value=value)
        if font:
            cel.font = font
        if fill:
            cel.fill = fill
        if alignment:
            cel.alignment = alignment
        if border:
            cel.border = border

    @staticmethod
    def _chave_puxada_exportacao(puxada: Puxada) -> Tuple:
        """Chave usada para juntar puxadas identicas no Excel."""
        return (
            int(puxada.largura_jumbo),
            tuple(int(bob.largura) for bob in puxada.bobinas),
            tuple(str(eixo) for eixo in puxada.eixos),
            int(puxada.refile_esquerdo_mm),
            int(puxada.refile_direito_mm),
            bool(puxada.completa_jumbo),
            str(puxada.faixa_refile),
        )

    @classmethod
    def _agrupar_puxadas_iguais_exportacao(
        cls, plano: List[Puxada]
    ) -> List[Tuple[Puxada, int]]:
        """Agrupa puxadas iguais, somando as repeticoes para exportacao."""
        grupos: List[Tuple[Puxada, int]] = []
        indice_por_chave: Dict[Tuple, int] = {}

        for puxada in plano:
            chave = cls._chave_puxada_exportacao(puxada)
            repeticao = int(puxada.repeticao)
            if chave in indice_por_chave:
                idx = indice_por_chave[chave]
                base, total = grupos[idx]
                grupos[idx] = (base, total + repeticao)
                continue

            indice_por_chave[chave] = len(grupos)
            grupos.append((puxada, repeticao))

        return grupos

    def _escrever_bloco_facas_excel(
        self,
        ws,
        puxada: Puxada,
        indice: int,
        repeticao_total: int,
        start_col: int,
        medida_inicial_mm: int,
        header_font,
        header_fill,
        yellow_fill,
        peach_fill,
        center,
        header_center,
        thin_border,
    ) -> None:
        facas, pos_final, total_corte = calcular_facas_puxada(
            puxada, medida_inicial_mm
        )
        col = start_col
        headers = ["FACA", "POSICAO", "BOBINAS", "Inferior", "Superior"]

        self._escrever_celula(
            ws, 1, col, f"PUXADA {indice:02d} ({repeticao_total}x)",
            font=header_font, fill=header_fill,
            alignment=header_center, border=thin_border,
        )
        self._escrever_celula(ws, 1, col + 1, "JUMBO", alignment=center, border=thin_border)
        self._escrever_celula(
            ws, 1, col + 2, puxada.largura_jumbo,
            fill=yellow_fill, alignment=center, border=thin_border,
        )
        self._escrever_celula(ws, 2, col + 1, "REFILE", alignment=center, border=thin_border)
        self._escrever_celula(
            ws, 2, col + 2, puxada.refile_esquerdo_mm,
            fill=yellow_fill, alignment=center, border=thin_border,
        )
        self._escrever_celula(ws, 3, col + 1, "MED. INIC.", alignment=center, border=thin_border)
        self._escrever_celula(
            ws, 3, col + 2, medida_inicial_mm,
            fill=yellow_fill, alignment=center, border=thin_border,
        )

        for offset, label in enumerate(headers):
            self._escrever_celula(
                ws, 4, col + offset, label,
                font=header_font, fill=header_fill,
                alignment=header_center, border=thin_border,
            )

        self._escrever_celula(
            ws, 5, col, "Refile esq", alignment=center, border=thin_border
        )
        self._escrever_celula(
            ws, 5, col + 2, puxada.refile_esquerdo_mm,
            fill=peach_fill, alignment=center, border=thin_border,
        )

        row = 6
        for idx, pos, largura, eixo in facas:
            self._escrever_celula(ws, row, col, idx, alignment=center, border=thin_border)
            self._escrever_celula(
                ws, row, col + 1, pos,
                fill=yellow_fill, alignment=center, border=thin_border,
            )
            self._escrever_celula(
                ws, row, col + 2, largura,
                fill=yellow_fill, alignment=center, border=thin_border,
            )
            if eixo == "Inferior":
                self._escrever_celula(
                    ws, row, col + 3, largura, alignment=center, border=thin_border
                )
                self._escrever_celula(ws, row, col + 4, "", alignment=center, border=thin_border)
            else:
                self._escrever_celula(ws, row, col + 3, "", alignment=center, border=thin_border)
                self._escrever_celula(
                    ws, row, col + 4, largura, alignment=center, border=thin_border
                )
            row += 1

        self._escrever_celula(
            ws, row, col, "Refile dir", alignment=center, border=thin_border
        )
        self._escrever_celula(
            ws, row, col + 1, pos_final,
            fill=yellow_fill, alignment=center, border=thin_border,
        )
        self._escrever_celula(
            ws, row, col + 2, puxada.refile_direito_mm,
            fill=peach_fill, alignment=center, border=thin_border,
        )
        row += 1
        self._escrever_celula(
            ws, row, col, "TOTAL DE CORTE", alignment=center, border=thin_border
        )
        self._escrever_celula(
            ws, row, col + 2, total_corte,
            fill=yellow_fill, alignment=center, border=thin_border,
        )

    # ---- dados iniciais ----------------------------------------------------

    def _carregar_dados_iniciais(self) -> None:
        try:
            self.itens_entrada = self.db.carregar_estoque()
            self.estoque_atividades = {
                int(l): int(q) for (l, q) in self.itens_entrada
            }
            self.itens_entrada = sorted(
                self.estoque_atividades.items(), key=lambda x: x[0], reverse=True
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar dados iniciais:\n{e}")

    def _atualizar_texto_estoque(self) -> None:
        try:
            self.lbl_estoque_titulo.grid(
                row=0, column=0, sticky="nw", padx=4, pady=(4, 8)
            )
            if not self.estoque_atividades and not self.itens_entrada:
                self.txt.delete("1.0", "end")
                return
            if self.estoque_atividades:
                self.itens_entrada = sorted(
                    self.estoque_atividades.items(), key=lambda x: x[0], reverse=True
                )
            self.txt.delete("1.0", "end")
            for l, q in self.itens_entrada:
                n = int(q)
                plural = "bobina" if n == 1 else "bobinas"
                self.txt.insert("end", f"✔ {n} {plural} de {int(l)}mm\n")
        except Exception as e:
            messagebox.showerror(
                "Erro", f"Falha ao atualizar texto do estoque:\n{e}"
            )

    # ---- loading / progresso -----------------------------------------------

    def _iniciar_loading(self, msg: str = "Processando...") -> None:
        self.btn_gerar.configure(state="disabled")
        self.progress_bar.set(0)
        self.lbl_status_sidebar.configure(text=msg)
        self.lbl_status_sidebar.place(
            relx=0.5, rely=0.86, anchor="center", relwidth=0.85
        )
        self.progress_bar.place(
            relx=0.5, rely=0.92, anchor="center", relwidth=0.8
        )

    def _parar_loading(self) -> None:
        self.btn_gerar.configure(state="normal")
        self.progress_bar.place_forget()
        self.lbl_status_sidebar.place_forget()
        self.lbl_status_sidebar.configure(text="")

    def _poll_progresso(self) -> None:
        try:
            while True:
                frac, msg = self._progress_queue.get_nowait()
                self.progress_bar.set(min(1.0, max(0.0, float(frac))))
                self.lbl_status_sidebar.configure(text=msg)
        except queue.Empty:
            pass
        if self.btn_gerar.cget("state") == "disabled":
            self.after(50, self._poll_progresso)

    # ---- opções e visualização ---------------------------------------------

    def _set_opcoes_puxada(
        self, plano: List[Puxada], residuais: List[Tuple[int, int]] = []
    ) -> None:
        try:
            if not plano:
                self.cmb_puxada.configure(values=["(sem puxadas)"])
                self.cmb_puxada.set("(sem puxadas)")
                self._redesenhar_canvas()
                return
            valores = [f"PUXADA {i}" for i in range(1, len(plano) + 1)]
            self.cmb_puxada.configure(values=valores)
            self.cmb_puxada.set(valores[0])
            self._redesenhar_canvas()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao atualizar lista de puxadas:\n{e}")

    def _on_select_puxada(self, _valor: str) -> None:
        self._redesenhar_canvas()

    def _redesenhar_canvas(self) -> None:
        try:
            if not self.plano_atual:
                desenhar_puxada(self.canvas, None)
                return

            sel = self.cmb_puxada.get()
            if not sel.startswith("PUXADA"):
                return
            idx = int(sel.split()[-1]) - 1
            if idx < 0 or idx >= len(self.plano_atual):
                desenhar_puxada(self.canvas, None)
                return

            desenhar_puxada(self.canvas, self.plano_atual[idx])
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao desenhar gráfico:\n{e}")


def main() -> None:
    """Ponto de entrada do aplicativo."""
    app = AppMRX()
    app.mainloop()
